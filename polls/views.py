# -*-coding:Utf-8 -*

from django.http import JsonResponse
from django.shortcuts import render, redirect, reverse
# from django.template.loader import render_to_string
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models import Sum
from django.contrib import messages
from django.forms import formset_factory
# from django.utils.text import slugify
# from django.conf import settings
# from django.contrib.auth.password_validation import validate_password
# from django.core.exceptions import ValidationError
# from django.core.files import File
# from django.forms import inlineformset_factory

# from django.urls import reverse_lazy, resolve
# from django.views.generic.edit import FormView
# from django.views.generic import DetailView
# from django.views.generic.list import ListView
# from django.core.exceptions import ValidationError

import openpyxl

import re

from .forms import (
    UserForm,
    UserBaseForm,
    UserCompForm,
    UploadFileForm,
    GroupDetail,
    EventDetail,
    CompanyForm,
    QuestionDetail,
)

# from .tools import define_password
from .tools import set_chart_data, init_event, user_is_admin, create_new_user

from .pollsmail import PollsMail

import json

from .models import (
    Company,
    Event,
    Question,
    Choice,
    UserVote,
    UserGroup,
    Result,
    Procuration,
    UserComp,
)


# =======================
#  User management views
# =======================


@user_passes_test(lambda u: u.is_superuser)
def new_user(request):
    # ================================================
    # View to create users - for development purposes
    # Superuser only (poor controls)
    # ================================================

    error = False

    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            if 'comp_slug' in request.session:
                comp_slug = request.session['comp_slug']
            else:
                comp_slug = ''
            if User.objects.filter(username=username):
                user_exists = True
            else:
                user_data = {
                    'username':  username,
                    'last_name': password,
                    'first_name': "",
                    'email': "",
                    'phone_num': "",
                    'is_admin': False,
                }
                new_user, usr_comp = create_new_user(comp_slug, user_data)
                # User.objects.create_user(username=username, password=password)
    else:
        form = UserForm()

    return render(request, "polls/sign_up.html", locals())


def login_user(request):
    error = False

    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            user = authenticate(username=username, password=password)
            if user:
                login(request, user)
                try:
                    request.session['comp_slug'] = user.usercomp.company.comp_slug
                except:
                    # user has no usercomp (superuser) : nothing to store
                    pass
                return redirect(reverse("polls:index"))
            else:
                error = True
    else:
        form = UserForm()

    return render(request, "polls/login.html", locals())


def logout_user(request):
    logout(request)
    request.session.flush()
    return redirect(reverse("polls:index"))


# =======================
#     Template views
# =======================


def index(request):
    """ Home page """
    if request.user.is_superuser:
        comp_list = Company.objects.filter()
        return render(request, "polls/index.html", locals())
    elif request.user.is_authenticated:
        return redirect("polls:company_home", comp_slug=request.user.usercomp.company.comp_slug)
    else:
        return render(request, "polls/index.html", locals())


@login_required
def company_home(request, comp_slug):
    """ Company's home page """
    # For superuser (no dedicated company), sets the session variable to the selected company
    if request.user.is_superuser:
        request.session['comp_slug'] = comp_slug
    # Display event list
    company = Company.get_company(comp_slug)
    next_event_list = Event.get_next_events(company)
    return render(request, "polls/company_home.html", locals())


@login_required
def event(request, comp_slug, event_slug):
    """ Event details page """

    # Define event context
    # company = Company.get_company(comp_slug)
    event = Event.get_event(comp_slug, event_slug)
    question_list = Question.get_question_list(event)
    nb_questions = len(question_list)

    # Check if connected user is part of the event and is authorized to vote
    user_can_vote = False
    if UserGroup.user_in_event(event_slug, request.user.usercomp):
        user_can_vote = True

        # Get user's proxy status
        proxy_list, user_proxy, user_proxy_list = Procuration.get_proxy_status(
            event_slug, request.user.usercomp
        )

    return render(request, "polls/event.html", locals())


@login_required
def question(request, comp_slug, event_slug, question_no):
    """ Questions details page 
        Manage information display """

    # Necessary info for the template
    # company = Company.get_company(comp_slug)
    event = Event.get_event(comp_slug, event_slug)
    evt_group_list = UserGroup.get_list(event_slug)
    question = Question.get_question(event, question_no)
    choice_list = Choice.get_choice_list(event_slug)
    last_question = False

    # Start event - occur when staff only used "Launch event" button
    if request.user.is_staff and not event.current:
        # Event can be launched only if total groups' weight == 100
        if (
            UserGroup.objects.filter(event=event).aggregate(Sum("weight"))[
                "weight__sum"
            ]
            != 100
        ):
            return redirect("polls:event", comp_slug=comp_slug, event_slug=event_slug)
        else:
            # Initialize users vote & results table
            init_event(event)

    # Gather user's info about the current question
    if not request.user.is_staff:
        user_vote = UserVote.get_user_vote(event_slug, request.user.usercomp, question_no)

    # Check if current question is the last one
    if question_no == len(Question.get_question_list(event)):
        last_question = True

    return render(request, "polls/question.html", locals())


@login_required
def results(request, comp_slug, event_slug):
    """ Results page """

    # company = Company.get_company(comp_slug)
    event = Event.get_event(comp_slug, event_slug)
    question_list = Question.get_question_list(event)
    nb_questions = len(question_list)

    return render(request, "polls/results.html", locals())


# =======================
#      Action views
# =======================


def get_chart_data(request):
    """ Gather and send information to build charts via ajax get request """

    comp_slug = request.GET["comp_slug"]
    event_slug = request.GET["event_slug"]
    question_no = int(request.GET["question_no"])

    event = Event.get_event(comp_slug, event_slug)
    evt_group_list = UserGroup.get_list(event_slug)

    data = set_chart_data(event, evt_group_list, question_no)

    return JsonResponse(data)


def vote(request, comp_slug, event_slug, question_no):
    """ Manage users' votes """

    choice_id = request.POST["choice"]
    user_vote = UserVote.set_vote(event_slug, request.user.usercomp, question_no, choice_id)

    data = {"success": "OK", "nb_votes": user_vote.nb_user_votes}

    return JsonResponse(data)


def set_proxy(request):
    """ Set proxyholder """

    user = User.objects.get(id=request.POST["user"])
    proxy = User.objects.get(id=request.POST["proxy"])
    event = Event.get_event(request.POST["comp_slug"], request.POST["event_slug"])

    Procuration.set_user_proxy(event, user, proxy)
    PollsMail(
        "ask_proxy",
        event,
        sender=[user.email],
        recipient_list=[proxy.email],
        user=user,
        proxy=proxy,
    )

    data = {
        "proxy_f_name": proxy.first_name,
        "proxy_l_name": proxy.last_name,
        "proxy": request.POST["proxy"],
    }

    return JsonResponse(data)


def accept_proxy(request):
    """ Accept proxy request """

    user = User.objects.get(id=request.POST["user"])
    # decode list sent in JSON format
    proxy_list = json.loads(request.POST["cancel_list"])

    event = Event.get_event(request.POST["comp_slug"], request.POST["event_slug"])

    for proxy_id in proxy_list:
        Procuration.confirm_proxy(event, user, int(proxy_id))
        PollsMail(
            "confirm_proxy", event, sender=[user.email], user=user, proxy_id=proxy_id
        )

    data = {"status": "Success"}

    return JsonResponse(data)


def cancel_proxy(request):
    """ Cancel or refuse proxy """

    event_slug = request.POST["event"]

    if request.POST["Action"] == "Refuse":
        proxy_list = request.POST.getlist("user_proxy")
        for proxy in proxy_list:
            Procuration.cancel_proxy(event_slug, int(proxy), request.user)
    else:
        Procuration.cancel_proxy(event_slug, request.user)

    return redirect("polls:event", event_slug=event_slug)


# ========================================
#      Administration views & actions
# ========================================

# Company settings management

@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_options(request, comp_slug):
    '''
        Manage Company options
    '''
    company = Company.get_company(comp_slug)
    comp_form = CompanyForm(request.POST or None, instance=company)

    if request.method == "POST":

        if comp_form.is_valid():
            comp_form.save()

    return render(request, "polls/adm_options.html", locals())


# Users management

@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_users(request, comp_slug):
    '''
        Manage users
    '''
    # Company's users list
    company = Company.get_company(comp_slug)
    user_list = UserComp.objects.order_by('user__last_name').filter(company=company)

    # Prepare form to upload users from file
    
    if company.use_groups: upload_form = UploadFileForm(initial={'use_groups': True})
    else: upload_form = UploadFileForm()

    data = {
        'user_list': user_list,
        'upload_form': upload_form
    }
    return render(request, "polls/adm_users.html", locals())


@login_required
def adm_user_profile(request, comp_slug, usr_id=0):
    '''User profile page'''
    access_admin = user_is_admin(comp_slug, request.user)

    if usr_id > 0:
        # A user exists : access only for the user himself of company admins
        profile_user = User.objects.get(pk=usr_id)
        if access_admin or profile_user == request.user:
            user_form = UserBaseForm(request.POST or None, instance=profile_user)
            usercomp_form = UserCompForm(request.POST or None, instance=profile_user.usercomp)
            user_form.fields['username'].disabled = True   # Disable updates of the field

        else:
            return redirect("polls:index")
    else:
        # User creation : company admins only
        if access_admin:
            user_form = UserBaseForm(request.POST or None)
            usercomp_form = UserCompForm(request.POST or None)
        else:
            return redirect("polls:index")

    if request.method == 'POST':
        if user_form.is_valid() and usercomp_form.is_valid():
            if usr_id == 0:
                # New user
                user_data = {
                    'username':  user_form.cleaned_data["username"],
                    'last_name': user_form.cleaned_data["last_name"],
                    'first_name': user_form.cleaned_data["first_name"],
                    'email': user_form.cleaned_data["email"],
                    'phone_num': usercomp_form.cleaned_data["phone_num"],
                    'is_admin': usercomp_form.cleaned_data["is_admin"],
                }
                new_user, usr_comp = create_new_user(comp_slug, user_data)
                msg = "Utilisateur {0} {1} créé avec succès".\
                    format(new_user.last_name, new_user.first_name)
                messages.success(request, msg)
            else:
                # Update user
                upd_user = user_form.save()
                usercomp_form.save()

                if upd_user == request.user:
                    msg = ("Votre profil a été modifié avec succès.")
                else:
                    msg = "Utilisateur {0} {1} modifié avec succès.".\
                        format(upd_user.last_name, upd_user.first_name)
        
                messages.success(request, msg)

            if usr_id == 0 or upd_user != request.user:
                # Update made at admin level : back to user list
                return redirect('polls:adm_users', comp_slug=comp_slug)
            else:
                # User's profile view : stay on the page
                return render(request, "polls/adm_user_profile.html", locals())

    return render(request, "polls/adm_user_profile.html", locals())


@login_required
def change_password(request, comp_slug):
    if request.method == 'POST':
        pwd_form = PasswordChangeForm(request.user, request.POST)
        if pwd_form.is_valid():
            user = pwd_form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Mot de passe modifié avec succès.')
            return redirect('polls:adm_user_profile', comp_slug=comp_slug, usr_id=request.user.id)
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        pwd_form = PasswordChangeForm(request.user)

    return render(request, "polls/change_pwd.html", locals())
        


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_delete_user(request, comp_slug, usr_id):
    del_usr = User.objects.get(pk=usr_id)
    msg = "Utilisateur {0} {1} supprimé.".\
            format(del_usr.last_name, del_usr.first_name)

    User.objects.get(pk=usr_id).delete()

    messages.success(request, msg)
    return redirect("polls:adm_users", comp_slug=comp_slug)


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_load_users(request, comp_slug):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']

            wb = openpyxl.load_workbook(f)
            sheet = request.POST['sheet']
            ws = wb[sheet]

            # Check the files has the required information
            elt_list = ['Nom', 'Prénom', 'Mail', 'Téléphone', 'Username']

            # Gather values in file's first row - supposed to be headers to retreive data
            header = ws[1]
            titles = [cell.value for cell in header]

            try:
                index_list = [titles.index(elt) for elt in elt_list]
            except:
                messages.error(request, "Chargement du fichier impossible : la structure ne correspond pas à ce qui est attendu")
                return redirect("polls:adm_users")

            nb_lines = 0
            nb_warn = 0
            nb_err = 0
            err_messages = []
            warn_messages = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                nb_lines += 1
                unique_user = True

                last_name = row[0]
                first_name = row[1]
                email = row[2]
                phone_num = row[3]
                if row[4] is not None:
                    username = row[4]
                else:
                    username = last_name.lower()

                # Data controls :
                # - user unicity : if username alreay exists;
                #       * for provided usernames, raise an error and reject line
                #       * else (username == last_name) : check first name and email
                #           if both are equal, reject row
                #           else : define new id with first-name chars
                # - phone numer format (warning only)

                if User.objects.filter(username=username).exists():
                    db_user = User.objects.get(username=username)
                    if (db_user.last_name == last_name and db_user.first_name == first_name \
                                                      and db_user.email == email) \
                        or (row[4] is not None):
                        # user already exists
                        nb_err += 1
                        err_messages.insert(0, "Utilisateur {0} {1} non créé : il existe déjà".format(last_name, first_name))
                        continue  # directly goes no next iteration

                    else:
                        unique_user = False
                        i = 0
                        while unique_user == False:
                            username += first_name.lower()[i:i+1]
                            i += 1
                            if User.objects.filter(username=username).exists():
                                # It's another user than before : needs to check unicity again
                                db_user = User.objects.get(username=username)
                                if (db_user.last_name == last_name and db_user.first_name == first_name \
                                                                and db_user.email == email) \
                                    or (row[4] is not None):
                                    # user already exists
                                    nb_err += 1
                                    err_messages.insert(0, "Utilisateur {0} {1} non créé : il existe déjà".format(last_name, first_name))
                                    break

                                elif i >= len(first_name.lower()):
                                    # Not able to automatically define a unique username
                                    nb_err += 1
                                    err_messages.insert(0, "Utilisateur {0} {1} non créé : impossible de définir un nom d'utilisateur unique".format(last_name, first_name))
                                    break
                            else:
                                unique_user = True

                # Out of while loop : check if a unique username has been found
                if unique_user is True:
                    if phone_num is not None and not re.match(r'^0[0-9]([ .-]?[0-9]{2}){4}$', str(phone_num)):
                        phone_num = ''    # integrate no phone number rather than invalid one
                        nb_warn += 1
                        warn_messages.insert(0, "Utilisateur {0} {1} : numéro de téléphone invalide".format(last_name, first_name))

                    # Create user & usercomp
                    new_user = User.objects.create_user(
                        username=username,
                        password=username,
                        last_name=last_name,
                        first_name=first_name,
                        email=email
                    )

                    company = Company.get_company(request.session['comp_slug'])
                    usr_comp = UserComp.create_usercomp(
                        user=new_user,
                        company=company,
                        phone_num=phone_num
                    )
            
            # Messages are created at the end to ensure having the right order
            msg = "{0} utilisateurs sur {1} correctement intégrés, {2} avertissement(s).".format((nb_lines - nb_err), nb_lines, nb_warn)
            messages.success(request, msg)

            for err_msg in err_messages:
                messages.error(request, err_msg)
            for warn_msg in warn_messages:
                messages.warning(request, warn_msg)

            return redirect("polls:adm_users", comp_slug=comp_slug)


# Groups management

@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_groups(request, comp_slug):
    '''
        Manage users groups
    '''
    # Variables set to be integrated in locals()
    company = Company.get_company(comp_slug)
    # user_list = UserComp.get_users_in_comp(comp_slug)
    # group_list = []
    group_list = UserGroup.objects.filter(company__comp_slug=comp_slug, hidden=False).order_by('group_name')

    return render(request, "polls/adm_groups.html", locals())


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_group_detail(request, comp_slug, grp_id=0):
    company = Company.get_company(comp_slug)
    if grp_id > 0:
        current_group = UserGroup.objects.get(id=grp_id)
        group_form = GroupDetail(request.POST or None, instance=current_group)
    else:
        group_form = GroupDetail(request.POST or None)
        group_form.fields['all_users'].queryset = UserComp.objects.\
                                                    filter(company=company).\
                                                    order_by('user__last_name', 'user__first_name')

    if request.method == 'POST':
        # Test if the list contains values and convert the string in a list of user IDs
        usr_list = []
        if request.POST.get('users_in_group', False):
            usr_list = [int(elt) for elt in request.POST['users_in_group'].split('-') if elt != ""]

        group_form.fields['users'].queryset = UserComp.objects.filter(id__in=usr_list).\
                                                        order_by('user__last_name', 'user__first_name')
        group_form.fields['all_users'].queryset = UserComp.objects.exclude(id__in=usr_list)

        if group_form.is_valid():
            if grp_id == 0:
                # Create empty group
                group_data = {
                    "company": company,
                    "group_name": group_form.cleaned_data["group_name"],
                    "weight": group_form.cleaned_data["weight"],
                }
                new_group = UserGroup.create_group(group_data, user_list=usr_list)
            else:
                # Update group
                new_group = group_form.save()

                # Remove all users then add the ones in the form's list
                group_usr_list = UserComp.objects.filter(usergroup=new_group)
                new_group.users.remove(*group_usr_list)
                new_group.users.add(*usr_list)
                new_group.save()

            # Update form according to latest changes
            group_form.fields['all_users'].queryset = UserComp.objects.\
                                                            exclude(id__in=usr_list).\
                                                            order_by('user__last_name', 'user__first_name')
            group_form.fields['users_in_group'].initial = "-".join([str(elt.id) for elt in new_group.users.all()])
        else:
            print("****** FORMULAIRE NON VALIDE *******")
            print(group_form.errors)

    return render(request, "polls/adm_group_detail.html", locals())


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_delete_group(request, comp_slug, grp_id):
    del_grp = UserGroup.objects.get(pk=grp_id)
    msg = "Groupe {0} supprimé.".format(del_grp.group_name)

    UserGroup.objects.get(pk=grp_id).delete()

    messages.success(request, msg)
    return redirect("polls:adm_groups", comp_slug=comp_slug)


# Event management

@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_events(request, comp_slug):
    '''
        Manage events
    '''
    # Define context variables
    company = Company.get_company(comp_slug)
    next_events = Event.get_next_events(company)
    old_events = Event.get_old_events(company)
    return render(request, "polls/adm_events.html", locals())


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_event_detail(request, comp_slug, evt_id=0):
    '''
        Manage events creation and options
    '''
    company = Company.get_company(comp_slug)
    # QuestionFormset = formset_factory(QuestionDetail, extra=3)

    if evt_id > 0:
        current_event = Event.objects.get(id=evt_id)
        event_form = EventDetail(request.POST or None, instance=current_event)
        event_form.fields['groups'].initial= current_event.groups.all()
        # question_set = QuestionFormset(request.POST, initial=list(Question.objects.filter(event=current_event)))

    else:
        event_form = EventDetail(request.POST or None)
        # question_set = QuestionFormset()

    event_form.fields['groups'].queryset= UserGroup.objects.\
                                            filter(company=company, hidden=False).\
                                            order_by('group_name')
    
    if request.method == 'POST':
        if event_form.is_valid():
        # if any(event_form.is_valid(), question_set.is_valid()):
            if evt_id == 0:
                # Create new event
                event_data = {
                    "company": company,
                    "groups": event_form.cleaned_data["groups"],
                    "event_name": event_form.cleaned_data["event_name"],
                    "event_date": event_form.cleaned_data["event_date"],
                    "quorum": event_form.cleaned_data["quorum"],
                    "rule":event_form.cleaned_data["rule"]
                }
                new_event = Event.create_event(event_data)
            else:
                new_event = event_form.save()
                new_event.groups.clear()
                new_event = event_form.save()
                new_event.groups.add(*event_form.cleaned_data['groups'])

            # for item in question_set:
            #     print(item)
            #     if item.cleaned_data:
            #         print(item.cleaned_data)
            #         Question.create_or_update(new_event, item.cleaned_data)
        else:
            print("****** FORMULAIRE NON VALIDE *******")
            print(event_form.errors)
            # print(question_set.errors)

    return render(request, "polls/adm_event_detail.html", locals())


@user_passes_test(lambda u: u.is_superuser or (u.id is not None and u.usercomp.is_admin))
def adm_delete_event(request, comp_slug, evt_id):
    del_evt = Event.objects.get(pk=evt_id)
    msg = "Evénement {0} du {1} supprimé.".format(del_evt.event_name, del_evt.event_date)

    Event.objects.filter(pk=evt_id).delete()

    messages.success(request, msg)
    return redirect("polls:adm_events", comp_slug=comp_slug)
